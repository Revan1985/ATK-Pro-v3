#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SOLUZIONE COMPLETA:
1. Carica file HTML italiano
2. Carica blocchi dall'XLSX (colonna Italiano)
3. Sostituisce ogni blocco nel HTML con {{BLOCK_N}}
4. Salva template HTML con placeholder
5. Genera file HTML per le lingue selezionate
6. Ricalcola automaticamente le lunghezze delle righe delle finestre box-drawing

USO:
  python generate_all_from_merged_xlsx.py               # Genera tutte le lingue (tranne IT)
  python generate_all_from_merged_xlsx.py en            # Genera solo inglese
  python generate_all_from_merged_xlsx.py en es fr      # Genera inglese, spagnolo e francese
"""

import openpyxl
import shutil
import os
import re
import sys

def find_windows(lines):
    """
    Trova tutte le finestre box-drawing nell'HTML.
    Gestisce finestre annidate identificando la finestra più ESTERNA.
    Ritorna lista di (start_idx, end_idx) per ogni finestra principale.
    """
    windows = []
    nesting_level = 0
    start_idx = None
    
    for idx, line in enumerate(lines):
        if '┌' in line:
            if nesting_level == 0:
                # Inizio di una nuova finestra principale
                start_idx = idx
            nesting_level += 1
        elif '┘' in line:
            nesting_level -= 1
            if nesting_level == 0 and start_idx is not None:
                # Fine della finestra principale
                windows.append((start_idx, idx))
                start_idx = None
    
    return windows

def get_max_content_length_in_window(lines, start, end):
    """
    Trova la lunghezza di caratteri (non visiva) massima nelle righe della finestra.
    Questa sarà la lunghezza target per TUTTE le righe.
    """
    max_length = 0
    
    for idx in range(start, end + 1):
        line = lines[idx]
        match = re.search(r'<span[^>]*>(.*?)</span>', line)
        if match:
            content = match.group(1)
            max_length = max(max_length, len(content))
    
    return max_length

def adjust_line_to_length(line, target_length):
    """
    Adatta una riga alla lunghezza target (in caratteri, non visiva).
    Per righe con emoji, usa MENO spazi perché l'emoji occupa 2 spazi visivi.
    """
    match = re.search(r'<span[^>]*>(.*?)(?:</span>|$)', line)
    if not match:
        return line
    
    content = match.group(1)
    
    # Gestisci righe con bordi verticali │
    if content.startswith('│') and content.endswith('│'):
        inner = content[1:-1].rstrip()  # Rimuovi bordi e spazi finali
        
        # Conta emoji nel contenuto (emoji occupa 2 spazi visivi ma 1 carattere)
        emoji_count = sum(1 for char in inner if ord(char) > 0x1F300)
        
        # Per mantenere allineamento visivo:
        # Se c'è un emoji, aggiungi 1 spazio in meno (perché l'emoji occupa 2 spazi visivi)
        char_count = len(inner)
        visual_spaces_occupied = char_count + emoji_count
        
        # Calcola spazi necessari
        spaces_needed = target_length - 2 - visual_spaces_occupied
        if spaces_needed < 0:
            spaces_needed = 0
            
        new_content = f"│{inner}{' ' * spaces_needed}│"
        return line.replace(content, new_content)
    
    # Gestisci righe con bordi orizzontali ┌─┐ └─┘ ├─┤
    elif any(char in content for char in ['┌', '┐', '└', '┘', '├', '┤']):
        new_content = content[0] + '─' * (target_length - 2) + content[-1]
        # Preserva tag di chiusura se presente
        if line.endswith('</span>'):
            return line.replace(content, new_content)
        else:
            # Riga spezzata senza </span>
            return line[:line.index('>')+1] + new_content
    
    return line

def find_all_windows_with_nesting(lines):
    """
    Trova TUTTE le finestre (esterne e innestate), restituendo per ognuna:
    (start_idx, end_idx, nesting_depth)
    Ordinate per profondità crescente (prima esterne depth=0, poi innestate depth=1, ecc.)
    """
    all_windows = []
    stack = []  # Stack di indici di inizio finestra
    
    for idx, line in enumerate(lines):
        if '┌' in line:
            stack.append(idx)
        elif '┘' in line:
            if stack:
                start_idx = stack.pop()
                nesting_depth = len(stack)  # Profondità al momento della chiusura
                all_windows.append((start_idx, idx, nesting_depth))
    
    # Ordina per profondità CRESCENTE (prima finestre esterne, poi innestate)
    all_windows.sort(key=lambda x: x[2])
    return all_windows

def process_html_windows(html_content):
    """
    Processa tutte le finestre box-drawing: trova la lunghezza massima
    necessaria per ogni finestra e allunga tutte le righe di conseguenza.
    Processa prima finestre INNESTATE, poi esterne (così le esterne si adattano).
    Ogni riga viene processata SOLO dalla SUA finestra diretta, non da quelle esterne.
    """
    lines = html_content.split('\n')
    all_windows = find_all_windows_with_nesting(lines)
    
    # Inverte ordine: prima innestate (depth maggiore), poi esterne
    all_windows_reversed = sorted(all_windows, key=lambda x: x[2], reverse=True)
    
    # Marca a quale finestra appartiene DIRETTAMENTE ogni riga (depth più alto)
    line_ownership = {}  # {line_idx: (start, end, depth)}
    for start, end, depth in all_windows_reversed:
        for idx in range(start, end + 1):
            if idx not in line_ownership:
                line_ownership[idx] = (start, end, depth)
    
    # Per ogni finestra, adatta SOLO le righe che le appartengono DIRETTAMENTE
    for start, end, depth in all_windows_reversed:
        # Calcola max_length SOLO per righe dirette
        max_length = 0
        
        for idx in range(start, end + 1):
            # Processa solo se questa riga appartiene DIRETTAMENTE a questa finestra
            if idx in line_ownership and line_ownership[idx] == (start, end, depth):
                match = re.search(r'<span[^>]*>(.*?)(?:</span>|$)', lines[idx])
                if match:
                    content = match.group(1)
                    max_length = max(max_length, len(content))
        
        # Adatta SOLO righe dirette
        for idx in range(start, end + 1):
            # Processa solo se questa riga appartiene DIRETTAMENTE a questa finestra
            if idx in line_ownership and line_ownership[idx] == (start, end, depth):
                if '<span' in lines[idx] and 'dir="ltr"' in lines[idx]:
                    lines[idx] = adjust_line_to_length(lines[idx], max_length)
    
    return '\n'.join(lines)

print("=" * 100)
print("GENERAZIONE FILE HTML DA XLSX ACCORPATO")
print("=" * 100)

# 1. Carica XLSX
wb = openpyxl.load_workbook('smartcat_translation_template.xlsx')
ws = wb.active

# Estrai header (nomi lingue)
headers = [cell.value for cell in ws[1]]
print(f"\n[OK] Trovate {len(headers)} lingue")

# 2. Estrai tutti i blocchi di testo (riga per riga)
blocks_by_lang = {}
for col_idx, lang in enumerate(headers):
    if not lang:
        continue
    blocks_by_lang[lang] = {}
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=col_idx + 1)
        if cell.value:
            block_num = row_idx - 1  # BLOCK_1 inizia dalla riga 2
            blocks_by_lang[lang][block_num] = cell.value


# Usa la colonna Italiano come chiave per i placeholder
it_blocks = blocks_by_lang.get('Italiano', {})
print(f"[OK] Estratti {len(it_blocks)} blocchi italiani")

# 3. Carica file HTML base (IT)
with open('assets/it/testuali/guida.html', 'r', encoding='utf-8') as f:
    html_template = f.read()

# 4. Sostituisci blocchi italiani con placeholder
sorted_it_blocks = sorted(it_blocks.items(), key=lambda x: len(x[1]), reverse=True)

replacements_made = 0
for block_num, it_text in sorted_it_blocks:
    placeholder = f'{{{{BLOCK_{block_num}}}}}'
    if it_text in html_template:
        html_template = html_template.replace(it_text, placeholder)
        replacements_made += 1

print(f"[OK] Sostituiti {replacements_made} blocchi IT con placeholder")

# 5. Salva template (opzionale, per debug)
with open('guida_template.html', 'w', encoding='utf-8') as f:
    f.write(html_template)
print(f"[OK] Template salvato: guida_template.html")

# 6. Determina quali lingue generare
# Mappa nome lingua -> codice ISO
lang_map = {
    'English': 'en', 'Español': 'es', 'Français': 'fr', 'Deutsch': 'de',
    'Português': 'pt', 'Русский': 'ru', 'إيطالي': 'ar', 'Nederlands': 'nl',
    'איטלקית': 'he', '日本語': 'ja', 'Svenska': 'sv', '中文': 'zh',
    'Română': 'ro', 'Ελληνικά': 'el', 'Polski': 'pl', 'Dansk': 'da',
    'Norsk': 'no', 'Türkçe': 'tr', 'Tiếng Việt': 'vi'
}

# Lingue RTL (Right-to-Left)
rtl_languages = {'ar', 'he'}

# Mappa inversa: codice -> nome lingua
code_to_lang = {v: k for k, v in lang_map.items()}


# Analizza argomenti da riga di comando
if len(sys.argv) > 1:
    # Lingue specifiche richieste
    requested_codes = [arg.lower() for arg in sys.argv[1:]]
    languages_to_generate = []
    for code in requested_codes:
        if code in code_to_lang:
            languages_to_generate.append(code_to_lang[code])
        elif code == 'it':
            languages_to_generate.append('Italiano')
        else:
            print(f"[!]  Codice '{code}' non riconosciuto - disponibili: {', '.join(sorted(list(lang_map.values()) + ['it']))}")
    if not languages_to_generate:
        print("Nessuna lingua valida specificata. Uscita.")
        sys.exit(1)
    print(f"\n[i] Modalità: Genera solo {len(languages_to_generate)} lingua/e: {', '.join(languages_to_generate)}")
else:
    # Default: tutte le lingue, ora inclusa anche Italiano
    languages_to_generate = [lang for lang in headers if lang]
    print(f"\n[i] Modalità: Genera TUTTE le {len(languages_to_generate)} lingue (incluso IT)")

print(f"\n{'=' * 100}")
print(f"GENERAZIONE {len(languages_to_generate)} FILE HTML")
print(f"{'=' * 100}\n")


for lang_name in languages_to_generate:
    if lang_name == 'Italiano':
        lang_code = 'it'
    else:
        lang_code = lang_map.get(lang_name)
    if not lang_code:
        print(f"[!]  Codice sconosciuto per {lang_name} - saltato")
        continue
    # Prendi i blocchi per questa lingua
    lang_blocks = blocks_by_lang.get(lang_name, {})
    # Parti dal template
    html_content = html_template
    # Sostituisci placeholder con traduzioni
    replacements = 0
    for block_num, translated_text in lang_blocks.items():
        placeholder = f'{{{{BLOCK_{block_num}}}}}'
        if placeholder in html_content:
            html_content = html_content.replace(placeholder, translated_text)
            replacements += 1
    # *** Imposta attributi lang e dir corretti ***
    direction = 'rtl' if lang_code in rtl_languages else 'ltr'
    html_content = html_content.replace(
        '<html lang="it" dir="ltr">',
        f'<html lang="{lang_code}" dir="{direction}">'
    )
    # *** NUOVO: Ricalcola lunghezze righe finestre ***
    html_content = process_html_windows(html_content)
    # Salva file
    output_path = f'assets/{lang_code}/testuali/guida.html'
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html_content)
    print(f"{lang_code.upper():2} ({lang_name:15}) - {replacements:3} sostituzioni -> {output_path}")


print(f"\n{'=' * 100}")
print(f"[OK] COMPLETATO! Generati {len(languages_to_generate)} file HTML")
print(f"{'=' * 100}")
