#!/usr/bin/env python3
"""Verify that the portal workbook mirrors the Markdown portal matrices."""

from __future__ import annotations

import re
import sys
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
DOCS_DIR = ROOT / "docs_generali"
WORKBOOK_PATH = DOCS_DIR / "Matrice_portali.xlsx"

SOURCES = [
    {
        "markdown": DOCS_DIR / "matrice_portali_esistenti_ATK-Pro.md",
        "sheet": "Matrice portali esistenti",
        "headers": [
            "Chiave",
            "Portale",
            "Area",
            "Metodo tecnico osservato",
            "Rischio manutenzione",
            "Stato legale operativo",
            "Prossimo passo",
        ],
    },
    {
        "markdown": DOCS_DIR / "matrice_portali_candidati_ATK-Pro.md",
        "sheet": "Matrice portali candidati",
        "headers": [
            "Chiave candidata",
            "Portale",
            "Area strategica",
            "Fonte ufficiale consultata",
            "Rilevanza genealogica",
            "Prima lettura tecnica/legale",
            "Decisione provvisoria",
        ],
    },
]


def normalize_cell(value: object) -> str:
    text = "" if value is None else str(value)
    text = text.replace("`", "")
    return re.sub(r"\s+", " ", text).strip()


def parse_markdown_row(line: str) -> list[str]:
    return [normalize_cell(cell) for cell in line.strip().strip("|").split("|")]


def is_separator_row(row: list[str]) -> bool:
    return all(re.fullmatch(r":?-{3,}:?", cell) for cell in row)


def read_markdown_table(path: Path, expected_headers: list[str]) -> list[list[str]]:
    expected = [normalize_cell(header) for header in expected_headers]
    lines = path.read_text(encoding="utf-8").splitlines()

    current: list[list[str]] = []
    for line in lines + [""]:
        if line.startswith("|"):
            current.append(parse_markdown_row(line))
            continue

        if current:
            header = current[0]
            if header == expected:
                rows = [header]
                rows.extend(row for row in current[1:] if not is_separator_row(row))
                return rows
            current = []

    raise ValueError(f"No matching Markdown table found in {path}")


def read_workbook_sheet(path: Path, sheet_name: str) -> list[list[str]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Missing sheet: {sheet_name}")

    sheet = workbook[sheet_name]
    rows: list[list[str]] = []
    for row in sheet.iter_rows(values_only=True):
        values = [normalize_cell(cell) for cell in row]
        if any(values):
            rows.append(values)
    workbook.close()
    return rows


def compare_rows(label: str, markdown_rows: list[list[str]], workbook_rows: list[list[str]]) -> list[str]:
    issues: list[str] = []

    if len(markdown_rows) != len(workbook_rows):
        issues.append(
            f"{label}: row count differs, Markdown={len(markdown_rows)} workbook={len(workbook_rows)}"
        )

    for index, (markdown_row, workbook_row) in enumerate(zip(markdown_rows, workbook_rows), start=1):
        if markdown_row != workbook_row:
            issues.append(
                f"{label}: row {index} differs\n"
                f"  Markdown: {markdown_row}\n"
                f"  Workbook: {workbook_row}"
            )

    if markdown_rows:
        keys = [row[0] for row in markdown_rows[1:]]
        duplicates = sorted({key for key in keys if keys.count(key) > 1})
        if duplicates:
            issues.append(f"{label}: duplicate keys in Markdown table: {', '.join(duplicates)}")

    return issues


def main() -> int:
    if not WORKBOOK_PATH.exists():
        print(f"ERROR: workbook not found: {WORKBOOK_PATH}", file=sys.stderr)
        return 1

    issues: list[str] = []
    summaries: list[str] = []

    for source in SOURCES:
        markdown_rows = read_markdown_table(source["markdown"], source["headers"])
        workbook_rows = read_workbook_sheet(WORKBOOK_PATH, source["sheet"])
        issues.extend(compare_rows(source["sheet"], markdown_rows, workbook_rows))
        summaries.append(f"{source['sheet']}: {len(markdown_rows) - 1} entries")

    if issues:
        print("Portal matrix workbook is out of sync:")
        for issue in issues:
            print(f"- {issue}")
        return 1

    print("Portal matrix workbook is aligned with Markdown sources.")
    for summary in summaries:
        print(f"- {summary}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
